import logging
import re
from pathlib import Path

from docx import Document as DocxDocument
from openpyxl import load_workbook
from PIL import Image

logger = logging.getLogger(__name__)

SUPPORTED_EXTENSIONS = {".pdf", ".docx", ".doc", ".xlsx", ".xls", ".txt", ".png", ".jpg", ".jpeg"}


class DocumentExtractor:
    """Extraction de contenu selon le format du fichier (SRP)."""

    def extract(self, file_path: str) -> str:
        path = Path(file_path)
        if not path.exists():
            raise FileNotFoundError(f"Fichier introuvable: {file_path}")

        extension = path.suffix.lower()
        if extension not in SUPPORTED_EXTENSIONS:
            return ""

        extractors = {
            ".pdf": self._extract_pdf,
            ".docx": self._extract_docx,
            ".doc": self._extract_docx,
            ".xlsx": self._extract_excel,
            ".xls": self._extract_excel,
            ".txt": self._extract_text,
            ".png": self._extract_image,
            ".jpg": self._extract_image,
            ".jpeg": self._extract_image,
        }

        extractor = extractors.get(extension)
        return extractor(path) if extractor else ""

    def _extract_pdf(self, path: Path) -> str:
        try:
            import fitz
        except ImportError as e:
            logger.warning("PyMuPDF / fitz non disponible: %s", e)
            return ""

        try:
            text_parts = []
            with fitz.open(path) as doc:
                for page in doc:
                    text_parts.append(page.get_text())
            return "\n".join(text_parts).strip()
        except Exception as e:
            logger.warning("Extraction PDF échouée: %s", e)
            return ""

    def _extract_docx(self, path: Path) -> str:
        doc = DocxDocument(path)
        return "\n".join(p.text for p in doc.paragraphs if p.text.strip())

    def _extract_excel(self, path: Path) -> str:
        wb = load_workbook(path, read_only=True, data_only=True)
        cells = []
        for sheet in wb.worksheets:
            for row in sheet.iter_rows(values_only=True):
                cells.extend(str(cell) for cell in row if cell is not None)
        wb.close()
        return " ".join(cells)

    def _extract_text(self, path: Path) -> str:
        return path.read_text(encoding="utf-8", errors="ignore")

    def _extract_image(self, path: Path) -> str:
        try:
            import pytesseract
            image = Image.open(path)
            return pytesseract.image_to_string(image, lang="fra+eng")
        except Exception as e:
            logger.warning("OCR indisponible: %s", e)
            return ""
